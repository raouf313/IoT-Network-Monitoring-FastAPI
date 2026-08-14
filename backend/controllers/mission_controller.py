"""
controllers/mission_controller.py — Logique métier des missions.
"""
import io
import unicodedata
import openpyxl
from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from config.database import get_db
from models.schemas import MissionCreate, MissionUpdate, MissionStatut
from services.tech_service import auto_assign_mission, auto_assign_all_available, redistribute_missions_of_unavailable_tech
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER

def get_missions() -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT m.*, u.nom technicien_nom, u.specialite tech_specialite,
               p.numero_ticket panne_ticket, p.type panne_type,
               p.description panne_description, p.priorite panne_priorite
        FROM missions m
        LEFT JOIN utilisateurs u ON m.technicien_id=u.id
        LEFT JOIN pannes p ON m.panne_id=p.id
        ORDER BY FIELD(m.priorite,'critique','eleve','moyen','faible'), m.date_creation DESC
    """)
    rows = cur.fetchall(); cur.close(); db.close()
    for r in rows:
        r["date_creation"] = r["date_creation"].isoformat()
        if r["date_fin"]: r["date_fin"] = r["date_fin"].isoformat()
    return {"missions": rows}


def creer_mission(data: MissionCreate) -> dict:
    if data.technicien_id:
        db = get_db(); cur = db.cursor(dictionary=True)
        cur.execute("SELECT disponible, nom FROM utilisateurs WHERE id=%s", (data.technicien_id,))
        tech = cur.fetchone(); cur.close(); db.close()
        if tech and not tech["disponible"]:
            raise HTTPException(400, f"Technicien {tech['nom']} est indisponible")
    db = get_db(); cur = db.cursor()
    cur.execute(
        "INSERT INTO missions (description,localisation,priorite,technicien_id,panne_id,statut) VALUES (%s,%s,%s,%s,%s,'en_attente')",
        (data.description, data.localisation, data.priorite, data.technicien_id, data.panne_id)
    )
    db.commit(); cur.close(); db.close()
    return {"status": "ok"}


def update_mission(mid: int, data: MissionUpdate) -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    if data.technicien_id is not None:
        cur.execute("SELECT disponible, nom FROM utilisateurs WHERE id=%s", (data.technicien_id,))
        tech = cur.fetchone()
        if tech and not tech["disponible"]:
            raise HTTPException(400, f"Technicien {tech['nom']} est indisponible")
    updates = []; values = []
    for k, v in data.dict(exclude_none=True).items():
        updates.append(f"{k}=%s"); values.append(v)
    if data.statut == "terminee":
        updates.append("date_fin=NOW()")
    if updates:
        cur2 = db.cursor()
        values.append(mid)
        cur2.execute(f"UPDATE missions SET {','.join(updates)} WHERE id=%s", values)
        db.commit(); cur2.close()
    cur.close(); db.close()
    return {"status": "ok"}


def update_mission_statut(mid: int, data: MissionStatut, add_background_task=None) -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    if data.statut == "terminee":
        # ── Vérification obligatoire : un rapport doit exister ──
        cur.execute("SELECT COUNT(*) as cnt FROM rapports WHERE mission_id=%s", (mid,))
        row = cur.fetchone()
        if not row or row["cnt"] == 0:
            cur.close(); db.close()
            raise HTTPException(
                400,
                "Impossible de terminer la mission : aucun rapport soumis. Veuillez d'abord créer un rapport."
            )
        cur.execute("UPDATE missions SET statut=%s, date_fin=NOW() WHERE id=%s", (data.statut, mid))
        db.commit()
        cur.execute("SELECT technicien_id FROM missions WHERE id=%s", (mid,))
        mission = cur.fetchone()
        tech_id = mission["technicien_id"] if mission else None
        cur.close(); db.close()
        if tech_id:
            db2 = get_db(); cur2 = db2.cursor()
            cur2.execute("UPDATE utilisateurs SET disponible=1, retour_disponible=NULL, duree_retour_minutes=NULL WHERE id=%s", (tech_id,))
            db2.commit(); cur2.close(); db2.close()
            auto_assign_mission(tech_id)
    elif data.statut == "refusee":
        cur.execute("SELECT technicien_id FROM missions WHERE id=%s", (mid,))
        mission = cur.fetchone()
        old_tech_id = mission["technicien_id"] if mission else None
        cur2 = db.cursor()
        cur.execute("""
            SELECT u.id, u.nom FROM utilisateurs u
            WHERE u.role='technicien' AND u.disponible=1 AND u.id != %s
              AND (SELECT COUNT(*) FROM missions m WHERE m.technicien_id=u.id
                   AND m.statut IN ('en_attente','acceptee','en_cours')) = 0
            ORDER BY (SELECT COUNT(*) FROM missions m2 WHERE m2.technicien_id=u.id AND m2.statut='terminee') ASC
            LIMIT 1
        """, (old_tech_id or 0,))
        nouveau_tech = cur.fetchone()
        if nouveau_tech:
            cur2.execute("UPDATE missions SET technicien_id=%s, statut='en_attente' WHERE id=%s", (nouveau_tech["id"], mid))
        else:
            cur2.execute("UPDATE missions SET technicien_id=NULL, statut='en_attente' WHERE id=%s", (mid,))
        if old_tech_id:
            cur2.execute("UPDATE utilisateurs SET disponible=1, retour_disponible=NULL, duree_retour_minutes=NULL WHERE id=%s", (old_tech_id,))
        db.commit(); cur.close(); cur2.close(); db.close()
    else:
        if data.statut in ('acceptee', 'en_cours'):
            # ── Vérifier que le technicien n'a pas déjà une mission active ──
            cur.execute("SELECT technicien_id FROM missions WHERE id=%s", (mid,))
            m = cur.fetchone()
            tech_id = m['technicien_id'] if m else None
            if tech_id:
                cur.execute(
                    "SELECT COUNT(*) v FROM missions WHERE technicien_id=%s AND statut IN ('acceptee','en_cours') AND id!=%s",
                    (tech_id, mid)
                )
                row = cur.fetchone()
                if row and row['v'] > 0:
                    cur.close(); db.close()
                    raise HTTPException(400, "Vous avez déjà une mission active. Terminez-la d'abord avant d'en accepter une nouvelle.")
                cur.execute("UPDATE utilisateurs SET disponible=0 WHERE id=%s", (tech_id,))
                db.commit()
        cur.execute("UPDATE missions SET statut=%s WHERE id=%s", (data.statut, mid))
        db.commit()
        cur.close(); db.close()
    return {"status": "ok"}


def terminer_mission(mid: int) -> dict:
    db = get_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id, technicien_id, statut FROM missions WHERE id=%s", (mid,))
        mission = cur.fetchone()
        if not mission: raise HTTPException(404, "Mission introuvable")
        
        # ── Vérification obligatoire : un rapport doit exister ──
        cur.execute("SELECT COUNT(*) as cnt FROM rapports WHERE mission_id=%s", (mid,))
        row = cur.fetchone()
        if not row or row["cnt"] == 0:
            raise HTTPException(
                400, 
                "Impossible de terminer la mission : aucun rapport soumis. Veuillez d'abord créer un rapport."
            )
        
        tech_id = mission["technicien_id"]
        cur.execute("UPDATE missions SET statut='terminee', date_fin=NOW() WHERE id=%s", (mid,))
        db.commit(); cur.close()
        nouvelle_mission = None
        if tech_id:
            cur2 = db.cursor()
            cur2.execute("UPDATE utilisateurs SET disponible=1, retour_disponible=NULL, duree_retour_minutes=NULL WHERE id=%s", (tech_id,))
            db.commit(); cur2.close()
            result = auto_assign_mission(tech_id)
            if result: nouvelle_mission = result[0]
        return {"status": "ok", "mission_terminee": mid, "nouvelle_mission": nouvelle_mission}
    finally:
        db.close()


def get_missions_technicien(tid: int) -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    cur.execute("SELECT disponible FROM utilisateurs WHERE id=%s", (tid,))
    tech = cur.fetchone()
    if tech and tech["disponible"]:
        cur.execute("SELECT COUNT(*) v FROM missions WHERE technicien_id=%s AND statut IN ('en_attente','acceptee','en_cours')", (tid,))
        if cur.fetchone()["v"] == 0:
            auto_assign_mission(tid)
    cur.close(); db.close()
    db2 = get_db(); cur2 = db2.cursor(dictionary=True)
    cur2.execute("""
        SELECT m.*, p.numero_ticket panne_ticket, p.type panne_type, p.description panne_description
        FROM missions m LEFT JOIN pannes p ON m.panne_id=p.id
        WHERE m.technicien_id=%s
        ORDER BY FIELD(m.statut,'en_cours','acceptee','en_attente','terminee','refusee'),
                 FIELD(m.priorite,'critique','eleve','moyen','faible')
    """, (tid,))
    rows = cur2.fetchall(); cur2.close(); db2.close()
    for r in rows:
        r["date_creation"] = r["date_creation"].isoformat()
        if r["date_fin"]: r["date_fin"] = r["date_fin"].isoformat()
    return {"missions": rows}


def assigner_mission(mid: int, data: dict) -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    
    # 1. Thabet fil-disponibilité mta3 el-technicien kima kont ta3mel
    cur.execute("SELECT disponible, nom FROM utilisateurs WHERE id=%s", (data["technicien_id"],))
    tech = cur.fetchone()
    if tech and not tech["disponible"]:
        raise HTTPException(400, f"Technicien {tech['nom']} est indisponible")
        
    cur2 = db.cursor()
    # 2. Update status el-mission (Traggé3ha 'en_cours' direct bech t-déclonchi l-affichage live)
    cur2.execute("UPDATE missions SET technicien_id=%s, statut='en_cours' WHERE id=%s", (data["technicien_id"], mid))
    
    # 3. 🔥 El-Ziyada el-wa7da hna: Bloki el-technicien fil-base bech ma3adch yokhroj disponible!
    cur2.execute("UPDATE utilisateurs SET disponible = 0 WHERE id = %s", (data["technicien_id"],))
    
    db.commit(); cur.close(); cur2.close(); db.close()
    return {"status": "ok"}

def signaler_blocage(mid: int, data: dict) -> dict:
    db = get_db(); cur = db.cursor(dictionary=True)
    cur.execute("SELECT technicien_id FROM missions WHERE id=%s", (mid,))
    mission = cur.fetchone()
    if not mission: raise HTTPException(404, "Mission introuvable")
    tech_id = mission["technicien_id"]
    cur2 = db.cursor()
    if tech_id:
        cur2.execute("UPDATE utilisateurs SET disponible=0 WHERE id=%s", (tech_id,))
    cur.execute("""
        SELECT u.id FROM utilisateurs u
        WHERE u.role='technicien' AND u.disponible=1 AND u.id != %s
          AND (SELECT COUNT(*) FROM missions m WHERE m.technicien_id=u.id
               AND m.statut IN ('en_attente','acceptee','en_cours')) = 0 LIMIT 1
    """, (tech_id or 0,))
    nouveau_tech = cur.fetchone()
    raison = data.get("raison", "Technicien bloqué")
    if nouveau_tech:
        cur2.execute("UPDATE missions SET technicien_id=%s, statut='en_attente' WHERE id=%s", (nouveau_tech["id"], mid))
        cur2.execute("INSERT INTO alertes (type,message,severite) VALUES ('warning',%s,'warning')", (f"⚠️ Mission #{mid} réassignée — {raison}",))
        db.commit(); cur.close(); cur2.close(); db.close()
        return {"status": "ok", "reassigne": True, "nouveau_technicien_id": nouveau_tech["id"]}
    else:
        cur2.execute("UPDATE missions SET technicien_id=NULL, statut='en_attente' WHERE id=%s", (mid,))
        cur2.execute("INSERT INTO alertes (type,message,severite) VALUES ('critique',%s,'critique')", (f"🔴 Mission #{mid} sans technicien — {raison}",))
        db.commit(); cur.close(); cur2.close(); db.close()
        return {"status": "ok", "reassigne": False}


def download_template_excel() -> StreamingResponse:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Missions"
    headers = ["description", "localisation", "priorite", "technicien_id"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill  = openpyxl.styles.PatternFill(fill_type="solid", fgColor="E30613")
        ws.column_dimensions[chr(64+col)].width = 25
    ws.append(["Vérification réseau fibre optique", "Tunis Centre", "eleve", ""])
    ws.append(["Maintenance équipements ADSL", "Sfax Station", "moyen", "2"])
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_missions_TT.xlsx"}
    )


async def import_missions_excel(file: UploadFile) -> dict:
    fname = file.filename or ""
    if not fname.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Format invalide. Utilisez .xlsx ou .xls")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content)); ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Impossible de lire le fichier Excel: {e}")
    # Normaliser les en-têtes : minuscules + suppression des accents (é→e, è→e, etc.)
    def _normalize(s):
        return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().strip()
    raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    headers = [_normalize(h.lower()) for h in raw_headers]
    missing = [r for r in ["description", "localisation", "priorite"] if r not in headers]
    if missing: raise HTTPException(400, f"Colonnes manquantes: {', '.join(missing)}. En-têtes trouvées: {headers}")
    db = get_db(); cur = db.cursor(); imported = 0; errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() for v in row): continue
        row_data = {h: str(v).strip() if v is not None else "" for h, v in zip(headers, row)}
        desc = row_data.get("description","").strip()
        if not desc: errors.append(f"Ligne {row_num}: description manquante"); continue
        prio = row_data.get("priorite","moyen").strip().lower()
        if prio not in ["faible","moyen","eleve","critique"]: prio = "moyen"
        tech_id = None
        tid_str = row_data.get("technicien_id","").strip()
        if tid_str and tid_str.replace('.','',1).isdigit():
            tid = int(float(tid_str))
            cur2 = db.cursor(dictionary=True)
            cur2.execute("SELECT id, disponible, nom FROM utilisateurs WHERE id=%s AND role='technicien'", (tid,))
            t = cur2.fetchone(); cur2.close()
            if t and t["disponible"]: tech_id = tid
        try:
            cur.execute(
                "INSERT INTO missions (description,localisation,priorite,technicien_id,statut) VALUES (%s,%s,%s,%s,'en_attente')",
                (desc, row_data.get("localisation","Non spécifié"), prio, tech_id)
            )
            imported += 1
        except Exception as e:
            errors.append(f"Ligne {row_num}: {e}")
    db.commit(); cur.close(); db.close()
    return {"status":"ok","imported":imported,"errors":errors,"message":f"✅ {imported} mission(s) importée(s)"}
def export_missions_pdf() -> StreamingResponse:
    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("""
        SELECT m.*, u.nom technicien_nom 
        FROM missions m 
        LEFT JOIN utilisateurs u ON m.technicien_id = u.id
    """)
    missions = cur.fetchall()
    cur.close()
    db.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#E30613'), alignment=TA_CENTER)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10)
    
    story.append(Paragraph("Tunisie Telecom — Liste Générale des Missions", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    table_data = [[Paragraph("<b>ID</b>", normal_style), Paragraph("<b>Description</b>", normal_style), Paragraph("<b>Statut</b>", normal_style)]]
    for m in missions:
        table_data.append([Paragraph(str(m['id']), normal_style), Paragraph(m['description'] or '', normal_style), Paragraph(m['statut'], normal_style)])
        
    t = Table(table_data, colWidths=[2*cm, 11*cm, 5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=missions.pdf"})